#!/usr/bin/env python3
from __future__ import annotations

import bisect
import json
import math
import re
import statistics
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA_XLSX = ROOT / "backtest-data.xlsx"
RESULTS_XLSX = ROOT / "backtest-results.xlsx"
OUTPUT_JS = ROOT / "dashboard-data.js"
METADATA_JSON = ROOT / "study-metadata.json"

TRADING_DAY_HORIZONS = {
    "1W": 5,
    "2W": 10,
    "1M": 21,
    "2M": 42,
    "3M": 63,
    "6M": 126,
    "9M": 189,
    "12M": 252,
}

CARD_HORIZONS = ["1W", "2W", "1M", "3M", "6M", "9M", "12M"]
GENERIC_PRICE_HEADERS = {"open", "high", "low", "close", "adj close", "adjusted close", "volume"}

HORIZON_ALIASES = {
    "1w": "1W",
    "1 wk": "1W",
    "1 week": "1W",
    "2w": "2W",
    "2 wk": "2W",
    "2 week": "2W",
    "2 weeks": "2W",
    "1m": "1M",
    "1 mo": "1M",
    "1 month": "1M",
    "2m": "2M",
    "2 mo": "2M",
    "2 month": "2M",
    "2 months": "2M",
    "3m": "3M",
    "3 mo": "3M",
    "3 month": "3M",
    "3 months": "3M",
    "6m": "6M",
    "6 mo": "6M",
    "6 month": "6M",
    "6 months": "6M",
    "9m": "9M",
    "9 mo": "9M",
    "9 month": "9M",
    "9 months": "9M",
    "12m": "12M",
    "12 mo": "12M",
    "12 month": "12M",
    "12 months": "12M",
    "1y": "12M",
    "1 yr": "12M",
    "1 year": "12M",
}

STAT_LABEL_ALIASES = {
    "median signal return": "Median Signal Return",
    "median signal returns": "Median Signal Return",
    "median signal": "Median Signal Return",
    "median return signal": "Median Signal Return",
    "signal median": "Median Signal Return",
    "average signal return": "Average Signal Return",
    "average signal returns": "Average Signal Return",
    "average signal": "Average Signal Return",
    "signal average": "Average Signal Return",
    "avg signal return": "Average Signal Return",
    "avg signal returns": "Average Signal Return",
    "avg signal": "Average Signal Return",
    "signal avg": "Average Signal Return",
    "mean signal return": "Average Signal Return",
    "mean signal returns": "Average Signal Return",
    "mean signal": "Average Signal Return",
    "signal mean": "Average Signal Return",
    "hit rate signal": "Signal Hit Rate",
    "signal hit rate": "Signal Hit Rate",
    "signal hit rates": "Signal Hit Rate",
    "signal z score": "Signal Z-Score",
    "z score": "Signal Z-Score",
    "zscore": "Signal Z-Score",
    "median all dataset return": "Median All-Dataset Return",
    "median all dataset returns": "Median All-Dataset Return",
    "median all dataset": "Median All-Dataset Return",
    "all dataset median": "Median All-Dataset Return",
    "median all day return": "Median All-Dataset Return",
    "median all day returns": "Median All-Dataset Return",
    "median all day": "Median All-Dataset Return",
    "all day median": "Median All-Dataset Return",
    "median baseline return": "Median All-Dataset Return",
    "median baseline returns": "Median All-Dataset Return",
    "baseline median": "Median All-Dataset Return",
    "all dataset median return": "Median All-Dataset Return",
    "all dataset median returns": "Median All-Dataset Return",
    "all day median return": "Median All-Dataset Return",
    "all day median returns": "Median All-Dataset Return",
    "baseline median return": "Median All-Dataset Return",
    "baseline median returns": "Median All-Dataset Return",
    "average all dataset return": "Average All-Dataset Return",
    "average all dataset returns": "Average All-Dataset Return",
    "average all dataset": "Average All-Dataset Return",
    "all dataset average": "Average All-Dataset Return",
    "avg all dataset return": "Average All-Dataset Return",
    "avg all dataset returns": "Average All-Dataset Return",
    "avg all dataset": "Average All-Dataset Return",
    "all dataset avg": "Average All-Dataset Return",
    "average all day return": "Average All-Dataset Return",
    "average all day returns": "Average All-Dataset Return",
    "average all day": "Average All-Dataset Return",
    "all day average": "Average All-Dataset Return",
    "avg all day return": "Average All-Dataset Return",
    "avg all day returns": "Average All-Dataset Return",
    "avg all day": "Average All-Dataset Return",
    "all day avg": "Average All-Dataset Return",
    "average baseline return": "Average All-Dataset Return",
    "average baseline returns": "Average All-Dataset Return",
    "average baseline": "Average All-Dataset Return",
    "all dataset average return": "Average All-Dataset Return",
    "all dataset average returns": "Average All-Dataset Return",
    "all dataset avg return": "Average All-Dataset Return",
    "all dataset avg returns": "Average All-Dataset Return",
    "all day average return": "Average All-Dataset Return",
    "all day average returns": "Average All-Dataset Return",
    "all day avg return": "Average All-Dataset Return",
    "all day avg returns": "Average All-Dataset Return",
    "baseline average return": "Average All-Dataset Return",
    "baseline average returns": "Average All-Dataset Return",
    "baseline average": "Average All-Dataset Return",
    "baseline avg return": "Average All-Dataset Return",
    "baseline avg returns": "Average All-Dataset Return",
    "baseline avg": "Average All-Dataset Return",
    "hit rate all dataset": "All-Dataset Hit Rate",
    "hit rates all dataset": "All-Dataset Hit Rate",
    "all dataset hit rate": "All-Dataset Hit Rate",
    "all dataset hit rates": "All-Dataset Hit Rate",
    "hit rate all day": "All-Dataset Hit Rate",
    "hit rates all day": "All-Dataset Hit Rate",
    "all day hit rate": "All-Dataset Hit Rate",
    "all day hit rates": "All-Dataset Hit Rate",
    "baseline hit rate": "All-Dataset Hit Rate",
    "baseline hit rates": "All-Dataset Hit Rate",
}


def read_study_metadata() -> dict[str, str]:
    if not METADATA_JSON.exists():
        return {}
    try:
        raw = json.loads(METADATA_JSON.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Could not parse {METADATA_JSON.name}: {exc}") from exc
    if not isinstance(raw, dict):
        raise ValueError(f"{METADATA_JSON.name} must contain a JSON object.")

    metadata = {}
    for field in ("title", "description", "slug"):
        value = raw.get(field)
        if isinstance(value, str) and value.strip():
            metadata[field] = value.strip()

    published_date = raw.get("publishedDate")
    if published_date not in (None, ""):
        if not isinstance(published_date, str):
            raise ValueError(f"{METADATA_JSON.name} publishedDate must be a date string.")
        parsed_published_date = parse_date(published_date)
        if not parsed_published_date:
            raise ValueError(f"{METADATA_JSON.name} publishedDate must be a recognized date.")
        metadata["publishedDate"] = parsed_published_date
    return metadata


def normalize_key(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("−", "-").replace("–", "-").replace("—", "-")
    text = re.sub(r"([a-zA-Z])([0-9])", r"\1 \2", text)
    text = re.sub(r"([0-9])([a-zA-Z])", r"\1 \2", text)
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    key = normalize_key(text)
    if not key:
        return ""
    if key == "signal date":
        return "Signal Date"
    if key in HORIZON_ALIASES:
        return HORIZON_ALIASES[key]
    if "max" in key and ("dd" in key or "drawdown" in key):
        if "12" in key or "1 y" in key or "1 year" in key:
            return "12M MaxDD"
    return re.sub(r"\s+", " ", text)


def clean_summary_line(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    text = re.sub(r"^[\u2022*\\-]\s*", "", text).strip()
    return text


def get_summary_section(sections: dict[str, list[str]], name: str, fuzzy: bool = True) -> list[str]:
    target = normalize_key(name)
    for section_name, lines in sections.items():
        key = normalize_key(section_name)
        if key == target or (fuzzy and (key.startswith(target) or target.startswith(key))):
            return lines
    return []


def canonical_stat_label(value: Any) -> str | None:
    key = normalize_key(value)
    return STAT_LABEL_ALIASES.get(key)


def find_canonical_stat_label(values: list[Any]) -> tuple[int, str, str] | None:
    for idx, value in enumerate(values):
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        canonical_label = canonical_stat_label(text)
        if canonical_label:
            return idx, canonical_label, text
    return None


def parse_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d, %Y", "%B %d, %Y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                pass
    return None


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isfinite(value):
            return float(value)
        return None
    if isinstance(value, str):
        stripped = value.strip().replace(",", "")
        if stripped.lower() in {"", "n/a", "na", "none", "-"}:
            return None
        is_percent = stripped.endswith("%")
        if is_percent:
            stripped = stripped[:-1]
        try:
            num = float(stripped)
            return num / 100 if is_percent else num
        except ValueError:
            return None
    return None


def serialize_cell(value: Any) -> Any:
    parsed_date = parse_date(value)
    if parsed_date:
        return parsed_date
    number = to_number(value)
    if number is not None:
        return number
    if value is None:
        return None
    return str(value)


def percentile(values: list[float], q: float) -> float | None:
    clean = sorted(v for v in values if v is not None and math.isfinite(v))
    if not clean:
        return None
    if len(clean) == 1:
        return clean[0]
    pos = (len(clean) - 1) * q
    lo = math.floor(pos)
    hi = math.ceil(pos)
    if lo == hi:
        return clean[lo]
    weight = pos - lo
    return clean[lo] * (1 - weight) + clean[hi] * weight


def median(values: list[float]) -> float | None:
    clean = [v for v in values if v is not None and math.isfinite(v)]
    return statistics.median(clean) if clean else None


def average(values: list[float]) -> float | None:
    clean = [v for v in values if v is not None and math.isfinite(v)]
    return sum(clean) / len(clean) if clean else None


def infer_trigger_name(results: dict[str, Any], criteria_description: str | None = None) -> str | None:
    parts = [criteria_description or "", results.get("title") or ""]
    parts.extend(str(item) for item in results.get("summaryText") or [])
    text = " ".join(part for part in parts if part)

    ratio = re.search(r"\b([A-Z]{2,8}/[A-Z]{2,8})\b", text)
    if ratio:
        return ratio.group(1)

    if re.search(r"\bRSI\b", text, re.IGNORECASE):
        return "RSI"

    dma = re.search(r"\b(\d{1,3})\s*(?:-| )?\s*(?:day\s*)?(?:moving\s+average|dma)\b", text, re.IGNORECASE)
    if dma:
        return f"{dma.group(1)}DMA"

    return None


def choose_source_columns(
    headers: list[str],
    sheet_name: str,
    preferred_asset_name: str | None = None,
    preferred_trigger_name: str | None = None,
) -> tuple[int, int | None, str, str]:
    keys = [normalize_key(header) for header in headers]
    preferred_asset_key = normalize_key(preferred_asset_name)
    preferred_trigger_key = normalize_key(preferred_trigger_name)
    asset_idx = None
    if preferred_asset_key:
        asset_idx = next((idx for idx, key in enumerate(keys) if key == preferred_asset_key), None)
    if asset_idx is None:
        asset_idx = next((idx for idx, key in enumerate(keys) if key in {"close", "adj close", "adjusted close"}), None)
    if asset_idx is None:
        asset_idx = 1

    asset_header = headers[asset_idx].strip() if asset_idx < len(headers) else ""
    asset_key = keys[asset_idx] if asset_idx < len(keys) else ""
    asset_name = sheet_name if asset_key in GENERIC_PRICE_HEADERS else asset_header
    if not asset_name:
        asset_name = sheet_name or "Asset"

    indicator_idx = None
    if preferred_trigger_key:
        indicator_idx = next((idx for idx, key in enumerate(keys) if idx != asset_idx and key == preferred_trigger_key), None)
    if indicator_idx is None:
        indicator_idx = next((idx for idx in range(1, len(keys)) if idx != asset_idx and keys[idx] and keys[idx] not in GENERIC_PRICE_HEADERS), None)
    if indicator_idx is None:
        indicator_idx = next((idx for idx in range(1, len(keys)) if idx != asset_idx and keys[idx]), None)
    indicator_name = headers[indicator_idx].strip() if indicator_idx is not None else ""
    return asset_idx, indicator_idx, asset_name, indicator_name


def read_source_data(preferred_asset_name: str | None = None, preferred_trigger_name: str | None = None) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(DATA_XLSX, data_only=True, read_only=True)
    if normalize_key(preferred_asset_name) in {"all", "signal", "signals", "cumulative"}:
        preferred_asset_name = None

    candidates: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        if len(headers) < 2:
            continue

        asset_idx, indicator_idx, asset_name, indicator_name = choose_source_columns(headers, sheet.title, preferred_asset_name, preferred_trigger_name)
        series: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_date = parse_date(row[0])
            asset = to_number(row[asset_idx]) if asset_idx < len(row) else None
            indicator = to_number(row[indicator_idx]) if indicator_idx is not None and indicator_idx < len(row) else None
            if row_date and asset is not None:
                series.append({"date": row_date, "asset": asset, "indicator": indicator})

        if series:
            candidates.append(
                {
                    "assetName": asset_name,
                    "indicatorName": indicator_name,
                    "series": series,
                    "preferredMatch": normalize_key(asset_name) == normalize_key(preferred_asset_name),
                }
            )

    if not candidates:
        raise ValueError("No usable rows found in backtest-data.xlsx.")

    selected = max(candidates, key=lambda candidate: (candidate["preferredMatch"], len(candidate["series"])))
    series = selected["series"]
    return {
        "assetName": selected["assetName"],
        "indicatorName": selected["indicatorName"],
        "series": series,
        "dateRange": {
            "start": series[0]["date"],
            "end": series[-1]["date"],
            "tradingDays": len(series),
        },
    }


def find_results_table(workbook: openpyxl.Workbook) -> tuple[openpyxl.worksheet.worksheet.Worksheet, int, int]:
    for sheet in workbook.worksheets:
        for row_idx in range(1, sheet.max_row + 1):
            row = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
            for col_idx, value in enumerate(row, start=1):
                if normalize_header(value) != "Signal Date":
                    continue
                headers = [normalize_header(cell) for cell in row[col_idx - 1 :]]
                horizon_count = sum(1 for header in headers if header in TRADING_DAY_HORIZONS)
                if horizon_count >= 2:
                    return sheet, row_idx, col_idx
    raise ValueError("Could not find a sheet with a 'Signal Date' header.")


def choose_results_start_col(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    signal_date_col: int,
) -> int:
    label_cols = [
        col
        for col in range(1, signal_date_col)
        if normalize_key(sheet.cell(header_row, col).value) in {"label", "row label", "metric", "stat", "statistic"}
    ]
    if label_cols:
        return label_cols[-1]

    previous_header = sheet.cell(header_row, signal_date_col - 1).value if signal_date_col > 1 else None
    if previous_header is not None and str(previous_header).strip():
        return signal_date_col - 1
    return signal_date_col


def find_row_type_col(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    header_start_col: int,
) -> int | None:
    return next(
        (
            col
            for col in range(1, header_start_col)
            if normalize_key(sheet.cell(header_row, col).value) in {"row type", "type"}
        ),
        None,
    )


def classify_result_row(row_type: Any, parsed_signal_date: str | None, label: str) -> str:
    row_type_key = normalize_key(row_type)
    if row_type_key in {"signal", "signals", "trade", "trades"}:
        return "signal"
    if row_type_key in {"summary", "stat", "stats", "statistic", "statistics"}:
        return "stat"
    if row_type_key in {"note", "notes"}:
        return "note"
    if parsed_signal_date:
        return "signal"
    if label.startswith(("Notes", "-", "*", "•")):
        return "note"
    return "stat"


def find_summary_sheet(workbook: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet | None:
    for sheet in workbook.worksheets:
        if normalize_key(sheet.title) == "summary":
            return sheet
    for sheet in workbook.worksheets:
        if normalize_key(sheet.title) in {"sumnary", "summry"}:
            return sheet
    return None


def row_text_cells(row: tuple[Any, ...]) -> list[list[str]]:
    cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
    if not cells:
        return []
    if len(cells) == 1:
        lines = [line.strip() for line in cells[0].splitlines() if line.strip()]
        return [[line] for line in lines]
    return [cells]


def starts_structured_summary_item(line: str) -> bool:
    return bool(re.match(r"^[A-Z0-9][^:]{0,42}:\s+", line))


def should_merge_summary_line(previous: str, current: str) -> bool:
    if not previous or not current:
        return False
    if starts_structured_summary_item(current):
        return False
    if previous.endswith((",", "—", "-", "–")):
        return True
    if previous.endswith((".", "!", "?", ")")):
        return False
    if current[0].islower():
        return True
    return True


def merge_wrapped_summary_lines(lines: list[str]) -> list[str]:
    merged: list[str] = []
    for line in lines:
        clean = re.sub(r"\s+", " ", clean_summary_line(line)).strip()
        if not clean:
            continue
        if merged and should_merge_summary_line(merged[-1], clean):
            merged[-1] = f"{merged[-1]} {clean}"
        else:
            merged.append(clean)
    return merged


def read_summary_text(workbook: openpyxl.Workbook) -> tuple[str, list[str], dict[str, list[str]]]:
    sheet = find_summary_sheet(workbook)
    if sheet is None:
        return "Backtest Visualizer", [], {}
    summary_rows: list[list[str]] = []
    text: list[str] = []
    for row in sheet.iter_rows(values_only=True):
        for cells in row_text_cells(row):
            summary_rows.append(cells)
            text.extend(cells)
    title = text[0] if text else "Backtest Visualizer"
    sections: dict[str, list[str]] = {}
    current_section: str | None = None
    for cells in summary_rows[1:]:
        line = cells[0]
        if re.match(r"^\s*[\u2022*\\-]\s+", line):
            if current_section:
                cleaned = clean_summary_line(line)
                if cleaned:
                    sections[current_section].append(cleaned)
            continue
        if len(cells) >= 2:
            current_section = line.strip().rstrip(":")
            if current_section:
                section = sections.setdefault(current_section, [])
                section.extend(cleaned for value in cells[1:] if (cleaned := clean_summary_line(value)))
            continue
        label_match = re.match(r"^([^:]{1,60}):\s*(.+)$", line)
        if label_match:
            current_section = label_match.group(1).strip()
            cleaned = clean_summary_line(label_match.group(2))
            if current_section:
                section = sections.setdefault(current_section, [])
                if cleaned:
                    section.append(cleaned)
            continue
        if current_section and (len(line) > 80 or re.search(r"[.!?]\s*$", line)):
            cleaned = clean_summary_line(line)
            if cleaned:
                sections.setdefault(current_section, []).append(cleaned)
            continue
        current_section = line.strip().rstrip(":")
        if current_section:
            sections.setdefault(current_section, [])
    sections = {name: merge_wrapped_summary_lines(lines) for name, lines in sections.items()}
    return title, text, sections


def extract_profile_asset_name(workbook: openpyxl.Workbook) -> str | None:
    patterns = [
        re.compile(r"[—–-]\s*([A-Za-z0-9./^& ]+?)\s+Forward\s+Returns?\b", re.IGNORECASE),
        re.compile(r"\b([A-Za-z0-9./^& ]+?)\s+Forward\s+Returns?\b", re.IGNORECASE),
    ]
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True):
            for value in row:
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                for pattern in patterns:
                    match = pattern.search(text)
                    if not match:
                        continue
                    candidate = re.sub(r"\s+", " ", match.group(1)).strip(" :-—–")
                    if candidate:
                        return candidate
    return None


def read_results() -> dict[str, Any]:
    workbook = openpyxl.load_workbook(RESULTS_XLSX, data_only=True, read_only=False)
    title, summary_text, summary_sections = read_summary_text(workbook)
    profile_asset_name = extract_profile_asset_name(workbook)
    sheet, header_row, signal_date_col = find_results_table(workbook)
    header_start_col = choose_results_start_col(sheet, header_row, signal_date_col)
    row_type_col = find_row_type_col(sheet, header_row, header_start_col)

    header_cols = [
        col
        for col in range(header_start_col, sheet.max_column + 1)
        if sheet.cell(header_row, col).value is not None and str(sheet.cell(header_row, col).value).strip()
    ]
    if not header_cols:
        raise ValueError("Could not read the Backtest Results header row.")
    last_header_col = max(header_cols)
    table_cols = list(range(header_start_col, last_header_col + 1))

    headers = [
        normalize_header(sheet.cell(header_row, col).value) or f"Column {col}"
        for col in table_cols
    ]
    try:
        signal_date_idx = headers.index("Signal Date")
    except ValueError as exc:
        raise ValueError("Could not find a 'Signal Date' column in the Backtest Results table.") from exc
    label_idx = next(
        (
            idx
            for idx, header in enumerate(headers)
            if normalize_key(header) in {"label", "row label", "metric", "stat", "statistic"}
        ),
        0,
    )

    table_rows: list[dict[str, Any]] = []
    signal_rows: list[dict[str, Any]] = []
    stats_rows: dict[str, dict[str, Any]] = {}
    notes: list[str] = []

    for row_num in range(header_row + 1, sheet.max_row + 1):
        values = [sheet.cell(row_num, col).value for col in table_cols]
        if not any(value is not None and str(value).strip() != "" for value in values):
            continue

        row_values = [serialize_cell(value) for value in values]
        parsed_signal_date = parse_date(values[signal_date_idx])
        label_text = "" if values[label_idx] is None else str(values[label_idx]).strip()
        first_text = "" if values[0] is None else str(values[0]).strip()
        stat_label_match = find_canonical_stat_label(values)
        stat_label_idx = stat_label_match[0] if stat_label_match else None
        detected_canonical_label = stat_label_match[1] if stat_label_match else None
        detected_label_text = stat_label_match[2] if stat_label_match else ""
        row_label = label_text or first_text or detected_label_text
        row_type = sheet.cell(row_num, row_type_col).value if row_type_col else None
        canonical_label = canonical_stat_label(row_label) or detected_canonical_label
        kind = classify_result_row(row_type, parsed_signal_date, row_label)
        if canonical_label:
            row_values[label_idx] = canonical_label
            if stat_label_idx is not None and stat_label_idx != label_idx:
                row_values[stat_label_idx] = None

        table_rows.append({"kind": kind, "label": canonical_label or row_label, "values": row_values})

        if kind == "signal" and parsed_signal_date:
            row_map = {
                headers[idx]: parsed_signal_date if idx == signal_date_idx else serialize_cell(values[idx])
                for idx in range(len(headers))
            }
            signal_rows.append({"date": parsed_signal_date, "values": row_map})
        elif kind == "stat" and (row_label or canonical_label):
            stats_key = canonical_label or row_label
            stats_rows[stats_key] = {
                headers[idx]: stats_key if idx == label_idx else row_values[idx]
                for idx in range(len(headers))
            }
        elif row_label:
            notes.append(row_label)

    horizons = [header for header in headers[1:] if header in TRADING_DAY_HORIZONS]

    return {
        "title": title,
        "profileAssetName": profile_asset_name,
        "summaryText": summary_text,
        "summarySections": summary_sections,
        "headers": headers,
        "horizons": horizons,
        "signalRows": signal_rows,
        "statsRows": stats_rows,
        "notes": notes,
        "resultTable": {"headers": headers, "rows": table_rows},
    }


def stat_value(stats_rows: dict[str, dict[str, Any]], row_name: str, horizon: str) -> float | None:
    row = stats_rows.get(row_name)
    if not row:
        return None
    return to_number(row.get(horizon))


def find_source_index(dates: list[str], signal_date: str) -> int | None:
    position = bisect.bisect_left(dates, signal_date)
    if position < len(dates) and dates[position] == signal_date:
        return position
    if position < len(dates):
        return position
    return None


def enrich_signals(source: dict[str, Any], results: dict[str, Any]) -> list[dict[str, Any]]:
    series = source["series"]
    dates = [point["date"] for point in series]
    enriched: list[dict[str, Any]] = []

    for signal in results["signalRows"]:
        entry_date = signal["values"].get("Entry Date")
        idx = find_source_index(dates, entry_date) if isinstance(entry_date, str) else None
        if idx is None:
            idx = find_source_index(dates, signal["date"])
        row_values = {
            key: value
            for key, value in signal["values"].items()
            if key != "Signal Date"
        }
        if idx is None:
            enriched.append(
                {
                    "date": signal["date"],
                    "asset": None,
                    "indicator": None,
                    "values": row_values,
                    "performance": [],
                    "completed12M": row_values.get("12M") is not None,
                }
            )
            continue
        start = series[idx]
        start_asset = start["asset"]
        performance = []
        for offset in range(0, TRADING_DAY_HORIZONS["12M"] + 1):
            current_idx = idx + offset
            if current_idx >= len(series):
                break
            point = series[current_idx]
            performance.append(
                {
                    "day": offset,
                    "date": point["date"],
                    "return": point["asset"] / start_asset - 1,
                }
            )

        enriched.append(
            {
                "date": signal["date"],
                "asset": start["asset"],
                "indicator": start["indicator"],
                "values": row_values,
                "performance": performance,
                "completed12M": row_values.get("12M") is not None,
            }
        )

    return enriched


def build_median_performance(signals: list[dict[str, Any]]) -> list[dict[str, Any]]:
    max_day = TRADING_DAY_HORIZONS["12M"]
    median_series = []
    for day in range(max_day + 1):
        values = []
        for signal in signals:
            if day < len(signal["performance"]):
                values.append(signal["performance"][day]["return"])
        day_median = median(values)
        if day_median is not None:
            median_series.append({"day": day, "return": day_median})
    return median_series


def build_comparison(results: dict[str, Any], signals: list[dict[str, Any]]) -> list[dict[str, Any]]:
    stats = results["statsRows"]
    comparison = []
    for horizon in results["horizons"]:
        sample_size = sum(1 for signal in signals if to_number(signal["values"].get(horizon)) is not None)
        comparison.append(
            {
                "horizon": horizon,
                "signalCount": sample_size,
                "signalAverage": stat_value(stats, "Average Signal Return", horizon),
                "signalMedian": stat_value(stats, "Median Signal Return", horizon),
                "allAverage": stat_value(stats, "Average All-Dataset Return", horizon),
                "allMedian": stat_value(stats, "Median All-Dataset Return", horizon),
                "signalHitRate": stat_value(stats, "Signal Hit Rate", horizon),
                "allHitRate": stat_value(stats, "All-Dataset Hit Rate", horizon),
                "zScore": stat_value(stats, "Signal Z-Score", horizon),
            }
        )
    return comparison


def build_distribution(signals: list[dict[str, Any]], horizons: list[str]) -> list[dict[str, Any]]:
    distribution = []
    for horizon in horizons:
        dots = []
        values = []
        for signal in signals:
            value = to_number(signal["values"].get(horizon))
            if value is None:
                continue
            values.append(value)
            dots.append(
                {
                    "date": signal["date"],
                    "value": value,
                    "asset": signal["asset"],
                    "indicator": signal["indicator"],
                }
            )
        clean = sorted(values)
        distribution.append(
            {
                "horizon": horizon,
                "values": dots,
                "box": {
                    "min": clean[0] if clean else None,
                    "q1": percentile(clean, 0.25),
                    "median": percentile(clean, 0.5),
                    "q3": percentile(clean, 0.75),
                    "max": clean[-1] if clean else None,
                },
            }
        )
    return distribution


def build_cards(signals: list[dict[str, Any]], results: dict[str, Any]) -> list[dict[str, Any]]:
    stats = results["statsRows"]
    completed_count = sum(1 for signal in signals if signal.get("completed12M"))
    cards = [
        {
            "kind": "count",
            "label": "Signal count",
            "value": len(signals),
            "detail": f"{completed_count} completed 12M windows",
        }
    ]

    horizon_samples: dict[str, list[float]] = {}
    card_horizons = [horizon for horizon in CARD_HORIZONS if horizon in results["horizons"]]
    for horizon in card_horizons:
        horizon_values = [to_number(signal["values"].get(horizon)) for signal in signals]
        clean = [value for value in horizon_values if value is not None]
        horizon_samples[horizon] = clean
        average_return = stat_value(stats, "Average Signal Return", horizon)
        if average_return is None:
            average_return = average(clean)
        cards.append(
            {
                "kind": "averageReturn",
                "label": f"{horizon} avg. return",
                "horizon": horizon,
                "value": average_return,
                "baseline": stat_value(stats, "Average All-Dataset Return", horizon),
                "sampleSize": len(clean),
            }
        )

    maxdd_values = [to_number(signal["values"].get("12M MaxDD")) for signal in signals]
    clean_dd = [value for value in maxdd_values if value is not None]
    cards.append(
        {
            "kind": "drawdown",
            "label": "Avg. max drawdown",
            "value": average(clean_dd),
            "baseline": stat_value(stats, "Average All-Dataset Return", "12M MaxDD"),
            "median": median(clean_dd),
            "sampleSize": len(clean_dd),
        }
    )

    for horizon in card_horizons:
        clean = horizon_samples[horizon]
        hit_rate = stat_value(stats, "Signal Hit Rate", horizon)
        if hit_rate is None and clean:
            hit_rate = sum(1 for value in clean if value > 0) / len(clean)
        cards.append(
            {
                "kind": "hitRate",
                "label": f"{horizon} hit rate",
                "horizon": horizon,
                "value": hit_rate,
                "baseline": stat_value(stats, "All-Dataset Hit Rate", horizon),
                "sampleSize": len(clean),
            }
        )
    return cards


def generate_description(source: dict[str, Any], signals: list[dict[str, Any]], results: dict[str, Any]) -> str:
    stats = results["statsRows"]
    signal_count = len(signals)
    completed = sum(1 for signal in signals if signal.get("completed12M"))
    one_month = stat_value(stats, "Average Signal Return", "1M")
    one_month_base = stat_value(stats, "Average All-Dataset Return", "1M")
    twelve_month = stat_value(stats, "Average Signal Return", "12M")
    twelve_month_base = stat_value(stats, "Average All-Dataset Return", "12M")
    twelve_median = stat_value(stats, "Median Signal Return", "12M")
    drawdown = average([to_number(signal["values"].get("12M MaxDD")) for signal in signals])

    short_term = "mixed"
    if one_month is not None and one_month_base is not None:
        short_term = "weaker than the all-day baseline" if one_month < one_month_base else "stronger than the all-day baseline"

    long_term = "mixed"
    if twelve_month is not None and twelve_month_base is not None:
        long_term = "above baseline" if twelve_month > twelve_month_base else "below baseline"

    parts = [
        f"This {source['assetName']} backtest found {signal_count} signals from {source['dateRange']['start']} through {source['dateRange']['end']}, with {completed} completed 12-month forward windows.",
        f"The setup's 1-month average forward return was {format_percent(one_month)} versus {format_percent(one_month_base)} for all days, making the short-term read {short_term}.",
        f"At 12 months, average signal return was {format_percent(twelve_month)} versus {format_percent(twelve_month_base)} baseline, while the median signal return was {format_percent(twelve_median)}.",
        f"Average 12-month max drawdown after signals was {format_percent(drawdown)}, so the sample points to meaningful path risk even when the 12-month endpoint is positive.",
    ]
    return " ".join(parts)


def dedupe_lines(lines: list[str]) -> list[str]:
    seen = set()
    clean_lines = []
    for line in lines:
        clean = re.sub(r"\s+", " ", clean_summary_line(line)).strip()
        if not clean:
            continue
        key = normalize_key(clean)
        if key in seen:
            continue
        seen.add(key)
        clean_lines.append(clean)
    return clean_lines


def collect_section_lines(
    sections: dict[str, list[str]],
    names: tuple[str, ...],
    limit_per_section: int | None = None,
    fuzzy: bool = True,
) -> list[str]:
    lines = []
    for name in names:
        section_lines = get_summary_section(sections, name, fuzzy=fuzzy)
        if limit_per_section is not None:
            section_lines = section_lines[:limit_per_section]
        lines.extend(section_lines)
    return dedupe_lines(lines)


def generate_summary_insights(results: dict[str, Any]) -> list[str]:
    sections = results.get("summarySections") or {}

    parts = []
    parts.extend(collect_section_lines(sections, ("Bottom line", "Takeaway"), 1))
    parts.extend(
        collect_section_lines(
            sections,
            ("Key findings", "Findings", "Headline trends", "Trends & Insights", "Results"),
            4,
        )
    )
    parts.extend(collect_section_lines(sections, ("Performance summary",), 5))
    parts.extend(collect_section_lines(sections, ("Statistical significance",), 2))
    parts.extend(
        collect_section_lines(
            sections,
            ("Notable individual outcomes", "Significant timeframes / notable observations", "Notable Timeframes"),
            2,
        )
    )
    return dedupe_lines(parts)[:6]


def generate_summary_description(results: dict[str, Any]) -> str | None:
    insights = generate_summary_insights(results)
    return " ".join(insights) if insights else None


def generate_criteria_details(results: dict[str, Any]) -> list[str]:
    sections = results.get("summarySections") or {}
    details = collect_section_lines(
        sections,
        ("Trigger criteria", "Trigger criteria plain English", "Methodology"),
        fuzzy=False,
    )
    if not details:
        details.extend(collect_section_lines(sections, ("Signal trigger", "Trigger"), 1))
    if details:
        timing = collect_section_lines(sections, ("Timing Convention",), 1)
        details.extend(timing)
    return dedupe_lines(details)


def generate_criteria_description(results: dict[str, Any]) -> str | None:
    details = generate_criteria_details(results)
    return " ".join(details) if details else None


def format_percent(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"{value * 100:.1f}%"


def build_payload() -> dict[str, Any]:
    metadata = read_study_metadata()
    results = read_results()
    criteria_details = generate_criteria_details(results)
    criteria_description = " ".join(criteria_details) if criteria_details else None
    trigger_name = infer_trigger_name(results, criteria_description)
    source = read_source_data(results.get("profileAssetName"), trigger_name)
    signals = enrich_signals(source, results)
    comparison = build_comparison(results, signals)
    distribution = build_distribution(signals, results["horizons"])
    cards = build_cards(signals, results)
    ai_description = generate_description(source, signals, results)
    summary_insights = generate_summary_insights(results)
    summary_description = " ".join(summary_insights) if summary_insights else None

    payload = {
        "title": metadata.get("title") or results["title"],
        "description": metadata.get("description") or ai_description,
        "slug": metadata.get("slug") or "",
        "publishedDate": metadata.get("publishedDate") or "",
        "assetName": source["assetName"],
        "indicatorName": source["indicatorName"],
        "triggerName": trigger_name or source["indicatorName"],
        "dateRange": source["dateRange"],
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "aiDescription": summary_description or ai_description,
        "criteriaDescription": criteria_description or "",
        "criteriaDetails": criteria_details,
        "summaryInsights": summary_insights,
        "summaryText": results["summaryText"],
        "summarySections": results.get("summarySections") or {},
        "horizons": results["horizons"],
        "cardHorizons": CARD_HORIZONS,
        "series": source["series"],
        "signals": signals,
        "medianPerformance": build_median_performance(signals),
        "comparison": comparison,
        "distribution": distribution,
        "cards": cards,
        "statsRows": results["statsRows"],
        "resultTable": results["resultTable"],
    }
    return payload


def main() -> None:
    payload = build_payload()
    body = json.dumps(payload, ensure_ascii=False, allow_nan=False, separators=(",", ":"))
    OUTPUT_JS.write_text(f"window.BACKTEST_DATA = {body};\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_JS}")


if __name__ == "__main__":
    main()
